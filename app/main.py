import os
import re
import uuid
import hashlib
from datetime import datetime, timezone
from io import BytesIO
from typing import List, Optional, Dict, Any, Tuple

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks, Query
from fastapi.middleware.cors import CORSMiddleware
from supabase import create_client
from openpyxl import load_workbook

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
SUPABASE_BUCKET = os.getenv("SUPABASE_BUCKET", "client-files")
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "*")

if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
    raise RuntimeError("Missing SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY env vars")

supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)

app = FastAPI(title="orcamento-backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in ALLOWED_ORIGINS.split(",")] if ALLOWED_ORIGINS != "*" else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------------
# Helpers
# -------------------------

def _sanitize_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r"[^\w\-. ]+", "_", name)
    name = re.sub(r"\s+", " ", name)
    return name[:120] if len(name) > 120 else name

def _to_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def _normalize_unit(u: str) -> str:
    u = _to_str(u).lower()
    u = u.replace(".", "").replace(" ", "")
    # comuns no seu cenário
    if u in ("m", "mt", "metro", "metros"):
        return "m"
    if u in ("pc", "pç", "peca", "peça", "pecas", "peças"):
        return "pç"
    if u in ("und", "un", "unid", "unidade", "unidades"):
        return "un"
    # fallback: mantém como veio, mas lower/trim
    return u or ""

def _parse_qty(q) -> Optional[float]:
    """
    Tenta converter quantidade do Excel para número.
    - aceita "1.234,56" / "1234,56" / "1234.56" / int/float do Excel
    """
    if q is None:
        return None
    if isinstance(q, (int, float)):
        return float(q)

    s = _to_str(q)
    if not s:
        return None

    # remove separador de milhar e normaliza decimal
    # caso "1.234,56"
    if re.match(r"^\d{1,3}(\.\d{3})+,\d+$", s):
        s = s.replace(".", "").replace(",", ".")
    # caso "1234,56"
    elif re.match(r"^\d+,\d+$", s):
        s = s.replace(",", ".")
    # remove lixo
    s = re.sub(r"[^\d\.\-]+", "", s)

    try:
        return float(s)
    except:
        return None

HEADER_HINTS_ITEM = ["item", "itens", "it.", "cód", "codigo", "code", "cod"]
HEADER_HINTS_DESC = ["descr", "descrição", "description", "material", "especifica", "specification"]
HEADER_HINTS_UNIT = ["und", "unid", "unidade", "unit"]
HEADER_HINTS_QTY  = ["qtd", "qtde", "quant", "quantidade", "qty", "quantity"]


def _detect_header_and_cols(ws, max_scan_rows: int = 60):
    """
    Retorna (header_row, item_col, desc_col, unit_col, qty_col)
    col_index é 1-based (openpyxl). Se não achar algum, retorna -1.
    """
    max_col = min(ws.max_column, 80)

    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            v = _to_str(ws.cell(row=r, column=c).value).lower()
            row_vals.append(v)

        # candidatos
        item_cols = [i+1 for i, v in enumerate(row_vals) if any(h in v for h in HEADER_HINTS_ITEM)]
        desc_cols = [i+1 for i, v in enumerate(row_vals) if any(h in v for h in HEADER_HINTS_DESC)]
        qty_cols  = [i+1 for i, v in enumerate(row_vals) if any(h in v for h in HEADER_HINTS_QTY)]
        unit_cols = [i+1 for i, v in enumerate(row_vals) if any(h in v for h in HEADER_HINTS_UNIT)]

        if not qty_cols:
            continue

        # tenta escolher desc_col de verdade
        desc_col = desc_cols[0] if desc_cols else -1
        qty_col = qty_cols[0]
        unit_col = unit_cols[0] if unit_cols else -1
        item_col = item_cols[0] if item_cols else -1

        # se desc_col não achou, tenta inferir: coluna com texto mais longo nas próximas linhas
        if desc_col == -1:
            best_c = -1
            best_len = 0
            for c in range(1, max_col + 1):
                total_len = 0
                samples = 0
                for rr in range(r+1, min(r+15, ws.max_row) + 1):
                    v = _to_str(ws.cell(row=rr, column=c).value)
                    if v:
                        total_len += len(v)
                        samples += 1
                avg_len = (total_len / samples) if samples else 0
                if avg_len > best_len:
                    best_len = avg_len
                    best_c = c
            desc_col = best_c if best_c != -1 else -1

        # proteção: desc_col não pode ser a mesma que item_col (quando o header é "Item")
        if desc_col == item_col and desc_col != -1:
            # tenta pegar outro candidato de descrição
            if len(desc_cols) > 1:
                desc_col = desc_cols[1]

        return (r, item_col, desc_col, unit_col, qty_col)

    return None

def _material_key(desc_final: str, unit_norm: str) -> str:
    # chave simples p/ MVP (depois vai evoluir)
    d = _to_str(desc_final).lower()
    d = re.sub(r"\s+", " ", d)
    return f"{d}|{unit_norm}"

# -------------------------
# Routes
# -------------------------

@app.get("/health")
def health():
    return {"ok": True}

@app.post("/runs/upload")
async def upload_run(
    files: List[UploadFile] = File(...),
    name: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    created_by: Optional[str] = Form(None),
):
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")

    run_id = str(uuid.uuid4())

    # cria o run
    try:
        supabase.table("runs").insert({
            "id": run_id,
            "name": name,
            "notes": notes,
            "created_by": created_by,
            "status": "queued",
        }).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB insert runs failed: {str(e)}")

    uploaded = []
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for f in files:
        original = f.filename or "arquivo.xlsx"
        original_s = _sanitize_filename(original)
        file_id = str(uuid.uuid4())

        content = await f.read()
        if not content:
            continue

        sha = hashlib.sha256(content).hexdigest()
        storage_path = f"uploads/{today}/{file_id}_{original_s}"

        # upload para o bucket privado
        try:
            supabase.storage.from_(SUPABASE_BUCKET).upload(
                path=storage_path,
                file=content,
                file_options={"content-type": f.content_type or "application/octet-stream"},
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Storage upload failed: {str(e)}")

        # registra no run_files
        try:
            supabase.table("run_files").insert({
                "id": file_id,
                "run_id": run_id,
                "bucket": SUPABASE_BUCKET,
                "storage_path": storage_path,
                "original_name": original,
                "sha256": sha,
                "size_bytes": len(content),
            }).execute()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"DB insert run_files failed: {str(e)}")

        uploaded.append({
            "file_id": file_id,
            "bucket": SUPABASE_BUCKET,
            "path": storage_path,
            "original_name": original,
            "sha256": sha,
            "size_bytes": len(content),
        })

    if not uploaded:
        raise HTTPException(status_code=400, detail="All files were empty")

    return {"run_id": run_id, "files": uploaded, "status": "queued"}


@app.get("/runs/{run_id}/status")
def run_status(run_id: str):
    run = supabase.table("runs").select("id,status,name,notes,error_message,created_at").eq("id", run_id).execute().data
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    run = run[0]

    # contagens (MVP)
    raw_count = supabase.table("raw_items").select("id", count="exact").eq("run_id", run_id).execute().count or 0
    syn_count = supabase.table("synthetic_items").select("id", count="exact").eq("run_id", run_id).execute().count or 0

    return {
        "run_id": run_id,
        "status": run["status"],
        "name": run.get("name"),
        "notes": run.get("notes"),
        "error_message": run.get("error_message"),
        "counts": {"raw": raw_count, "synthetic": syn_count},
        "created_at": run.get("created_at"),
    }


@app.get("/runs/{run_id}/synthetic")
def get_synthetic(
    run_id: str,
    limit: int = Query(200, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    q: Optional[str] = None,
):
    query = supabase.table("synthetic_items").select("id,desc_final,unit_norm,qty_total,sources,created_at").eq("run_id", run_id)
    if q:
        # ilike no desc_final (p/ MVP)
        query = query.ilike("desc_final", f"%{q}%")

    data = query.order("qty_total", desc=True).range(offset, offset + limit - 1).execute().data or []
    return {"run_id": run_id, "items": data, "limit": limit, "offset": offset}


def _process_run_background(run_id: str):
    """
    PROCESSAMENTO MVP:
    - baixa os arquivos do bucket
    - extrai linhas com heurística de cabeçalho
    - grava raw_items
    - gera synthetic_items agrupando por desc_raw + unidade (normalizada)
    """
    try:
        # evita reprocessamento concorrente
        run = supabase.table("runs").select("status").eq("id", run_id).execute().data
        if not run:
            return
        if run[0]["status"] == "processing":
            return

        supabase.table("runs").update({"status": "processing", "error_message": None}).eq("id", run_id).execute()

        # limpa dados antigos (reprocess)
        supabase.table("raw_items").delete().eq("run_id", run_id).execute()
        supabase.table("synthetic_items").delete().eq("run_id", run_id).execute()

        files = supabase.table("run_files").select("id,bucket,storage_path,original_name").eq("run_id", run_id).execute().data or []
        if not files:
            supabase.table("runs").update({"status": "error", "error_message": "No files in run_files"}).eq("id", run_id).execute()
            return

        # agregador do sintético
        agg: Dict[str, Dict[str, Any]] = {}
        raw_batch: List[Dict[str, Any]] = []

        for f in files:
            file_id = f["id"]
            bucket = f["bucket"]
            path = f["storage_path"]
            original_name = f["original_name"]

            content = supabase.storage.from_(bucket).download(path)
            wb = load_workbook(BytesIO(content), data_only=True, read_only=True)

            for ws in wb.worksheets:
                header = _detect_header_and_cols(ws)
                if not header:
                    continue  # MVP: ignora abas sem cabeçalho identificado

                header_row, item_col, desc_col, unit_col, qty_col = header

                for r in range(header_row + 1, ws.max_row + 1):
                    item_code = _to_str(ws.cell(row=r, column=item_col).value) if item_col != -1 else ""
                    desc_raw = _to_str(ws.cell(row=r, column=desc_col).value) if desc_col != -1 else ""
                    unit_raw = _to_str(ws.cell(row=r, column=unit_col).value) if unit_col != -1 else ""
                    qty_raw  = _to_str(ws.cell(row=r, column=qty_col).value)

                    qty_num = _parse_qty(qty_raw)
                    unit_norm = _normalize_unit(unit_raw)

                    def _has_letter(s: str) -> bool:
                        return bool(re.search(r"[A-Za-zÀ-ÿ]", s or ""))

                    include = bool(desc_raw) and _has_letter(desc_raw) and (qty_num is not None) and (qty_num > 0)

                    raw_row = {
                        "run_id": run_id,
                        "file_id": file_id,
                        "sheet_name": ws.title,
                        "row_number": r,
                        "desc_raw": desc_raw,
                        "unit_raw": unit_raw,
                        "qty_raw": qty_raw,
                        "qty_num": qty_num,
                        "include_in_synthetic": include,
                    }
                    raw_batch.append(raw_row)

                    if include:
                        # no MVP, desc_final = desc_raw (depois entra motor de regras)
                        desc_final = desc_raw
                        key = _material_key(desc_final, unit_norm)

                        if key not in agg:
                            agg[key] = {
                                "run_id": run_id,
                                "desc_final": desc_final,
                                "unit_norm": unit_norm,
                                "qty_total": 0.0,
                                "sources": [],
                            }
                        agg[key]["qty_total"] += float(qty_num)
                        agg[key]["sources"].append({
                            "file": original_name,
                            "sheet": ws.title,
                            "row": r,
                        })

                    # flush batch
                    if len(raw_batch) >= 500:
                        supabase.table("raw_items").insert(raw_batch).execute()
                        raw_batch = []

            # fecha workbook
            wb.close()

        # flush restante do raw
        if raw_batch:
            supabase.table("raw_items").insert(raw_batch).execute()

        # grava sintético em batch
        syn_rows = list(agg.values())
        for i in range(0, len(syn_rows), 200):
            supabase.table("synthetic_items").insert(syn_rows[i:i+200]).execute()

        supabase.table("runs").update({"status": "done"}).eq("id", run_id).execute()

    except Exception as e:
        supabase.table("runs").update({"status": "error", "error_message": str(e)}).eq("id", run_id).execute()


@app.post("/runs/{run_id}/process")
def process_run(run_id: str, background: BackgroundTasks):
    run = supabase.table("runs").select("id,status").eq("id", run_id).execute().data
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    status = run[0]["status"]
    if status == "processing":
        return {"run_id": run_id, "status": "processing", "message": "Already processing"}

    # dispara em background (assíncrono)
    background.add_task(_process_run_background, run_id)
    return {"run_id": run_id, "status": "processing", "message": "Processing started"}
